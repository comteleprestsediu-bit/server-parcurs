from flask import Flask, request, jsonify, send_file, session, redirect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

app = Flask(__name__)
app.secret_key = "parcurs_secret"

FILE = "curse.xlsx"

# 🔹 USERS
USERS = {
    "admin": "1234",
    "sofer": "1234"
}

# 🔹 INIT EXCEL (CU TITLU + STIL)
def init_excel():
    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active

        # 🔵 TITLU
        ws.merge_cells("A1:E1")
        title = ws["A1"]
        title.value = "Foaie de parcurs Flota Comteleprest"
        title.font = Font(size=14, bold=True)
        title.alignment = Alignment(horizontal="center")

        # 🔵 HEADER (rand 2)
        headers = [
            "nr_auto",
            "data",
            "locatie_plecare",
            "locatie_sosire",
            "km_parcurs"
        ]

        ws.append(headers)

        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

        # 📏 latime coloane
        widths = [15, 15, 40, 40, 15]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        wb.save(FILE)

init_excel()


# 🔹 CITESTE CURSE
def citeste_curse():
    wb = load_workbook(FILE)
    ws = wb.active

    curse = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue

        curse.append({
            "id": i - 2,
            "nr_auto": row[0] or "",
            "data": row[1] or "",
            "locatie_plecare": row[2] or "",
            "locatie_sosire": row[3] or "",
            "km_parcurs": str(row[4] or "0")
        })

    return curse


# 🔐 LOGIN
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = request.form.get("user")
        parola = request.form.get("parola")

        if USERS.get(user) == parola:
            session["user"] = user
            return redirect("/")
        else:
            return "Login greșit"

    return """
    <html>
    <head>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body class="bg-light">

    <div class="container mt-5">
        <div class="card p-4 shadow" style="max-width:400px;margin:auto;">
            <h3 class="mb-3">Login</h3>

            <form method="post">
                <input name="user" class="form-control mb-2" placeholder="User">
                <input name="parola" type="password" class="form-control mb-2" placeholder="Parola">
                <button class="btn btn-primary w-100">Login</button>
            </form>
        </div>
    </div>

    </body>
    </html>
    """


# 🔓 LOGOUT
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# 🔥 PAGINA PRINCIPALA
@app.route("/")
def index():
    if "user" not in session:
        return redirect("/login")

    curse = citeste_curse()

    html = """
    <html>
    <head>
        <title>Foaie Parcurs</title>
        <meta name="viewport" content="width=device-width, initial-scale=1">

        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">

        <style>
            body { background:#f5f7fa; }
            .card { border-radius:15px; }
        </style>
    </head>

    <body>

    <div class="container mt-4">

        <div class="card shadow p-4">

            <div class="d-flex justify-content-between">
                <h3>🚗 Raport curse</h3>
                <a href="/logout" class="btn btn-secondary">Logout</a>
            </div>

            <a href="/download" class="btn btn-success mt-2 mb-3">⬇️ Descarcă Excel</a>

            <input type="text" id="search" class="form-control mb-3" placeholder="🔍 Caută...">

            <div class="table-responsive">
                <table class="table table-striped table-hover" id="tabel">
                    <thead class="table-dark">
                        <tr>
                            <th>ID</th>
                            <th>Nr Auto</th>
                            <th>Data</th>
                            <th>Plecare</th>
                            <th>Sosire</th>
                            <th>Km</th>
                            <th>Șterge</th>
                        </tr>
                    </thead>
                    <tbody>
    """

    for c in curse:
        html += f"""
        <tr>
            <td>{c['id']}</td>
            <td>{c['nr_auto']}</td>
            <td>{c['data']}</td>
            <td>{c['locatie_plecare']}</td>
            <td>{c['locatie_sosire']}</td>
            <td>{c['km_parcurs']}</td>
            <td>
                <button class="btn btn-danger btn-sm" onclick="sterge({c['id']})">🗑</button>
            </td>
        </tr>
        """

    html += """
                    </tbody>
                </table>
            </div>

            <h5 id="total" class="mt-3"></h5>

        </div>
    </div>

    <script>
        document.getElementById("search").addEventListener("keyup", function() {
            let filter = this.value.toLowerCase();
            let rows = document.querySelectorAll("#tabel tbody tr");

            rows.forEach(row => {
                let text = row.innerText.toLowerCase();
                row.style.display = text.includes(filter) ? "" : "none";
            });
        });

        let total = 0;
        document.querySelectorAll("#tabel tbody tr").forEach(row => {
            let km = row.cells[5].innerText.replace(",", ".");
            total += parseFloat(km) || 0;
        });

        document.getElementById("total").innerText = "Total km: " + total.toFixed(2);

        function sterge(id) {
            if (!confirm("Ștergi cursa?")) return;

            fetch("/sterge_cursa/" + id, {
                method: "DELETE"
            })
            .then(res => res.json())
            .then(() => location.reload());
        }
    </script>

    </body>
    </html>
    """

    return html


# 🔹 DOWNLOAD
@app.route("/download")
def download_excel():
    return send_file(FILE, as_attachment=True)


# 🔹 API LISTA
@app.route("/api/curse")
def api_curse():
    return jsonify(citeste_curse())


# 🔹 ADAUGARE CURSA (FIX TELEFON)
@app.route("/adauga_cursa", methods=["POST"])
def adauga_cursa():
    try:
        data = request.get_json(silent=True)

        if not data:
            data = request.form

        nr_auto = data.get("nrAuto") or data.get("nr_auto", "")
        data_cursa = data.get("data", "")
        plecare = data.get("locatiePlecare") or data.get("locatie_plecare", "")
        sosire = data.get("locatieSosire") or data.get("locatie_sosire", "")
        km = data.get("kmParcurs") or data.get("km_parcurs", "")

        wb = load_workbook(FILE)
        ws = wb.active

        ws.append([nr_auto, data_cursa, plecare, sosire, km])

        wb.save(FILE)

        return {"status": "ok"}

    except Exception as e:
        print("EROARE:", e)
        return {"status": "error"}


# 🔥 STERGERE
@app.route("/sterge_cursa/<int:id>", methods=["DELETE"])
def sterge_cursa(id):
    try:
        wb = load_workbook(FILE)
        ws = wb.active

        ws.delete_rows(id + 3)

        wb.save(FILE)

        return {"status": "deleted"}

    except Exception as e:
        print("EROARE ȘTERGERE:", e)
        return {"status": "error"}


# 🔹 START
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
